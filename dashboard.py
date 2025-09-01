import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
import numpy as np
import openpyxl 
import io

# --- Configuração da Página e Estilo ---
st.set_page_config(
    page_title="Dashboard de Análise de Faturas",
    page_icon="🔵",
    layout="wide"
)

# --- INSTRUÇÃO IMPORTANTE: As credenciais agora são gerenciadas via st.secrets ---
# Certifique-se de ter o arquivo .streamlit/secrets.toml configurado
FILE_PATH = st.secrets["file_credentials"]["path"]
SHEET_NAME = st.secrets["file_credentials"]["sheet_name"]
SHEET_PASSWORD = st.secrets["file_credentials"]["password"]

# --- Funções Auxiliares ---
def format_currency(value):
    if pd.isna(value): return "R$ 0,00"
    return f"R$ {value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

def format_foreign_currency(value):
    if pd.isna(value): return "0.00"
    return f"{value:,.2f}"

# --- Funções de Carregamento e Tratamento de Dados ---
@st.cache_data(ttl=600)
def load_data(file_path, sheet_name, password):
    try:
        with open(file_path, "rb") as f:
            in_memory_file = io.BytesIO(f.read())
        workbook = openpyxl.load_workbook(in_memory_file)
        if sheet_name not in workbook.sheetnames:
            st.error(f"Erro: A aba '{sheet_name}' não foi encontrada na planilha. Abas disponíveis: {workbook.sheetnames}")
            return None, None
        sheet = workbook[sheet_name]
        if sheet.protection.sheet and password:
            sheet.protection.password = password
        output_buffer = io.BytesIO()
        workbook.save(output_buffer)
        output_buffer.seek(0)
        df = pd.read_excel(output_buffer, sheet_name=sheet_name, skiprows=1)
        df.rename(columns={'Data de Pagamento': 'data_pagamento', 'Valor em R$': 'valor_principal_rs','Número do RM': 'numero_rm', 'Correspondente': 'correspondente','ND recebida': 'nd_recebida', 'IOF': 'iof', 'TX CONTRATO': 'tx_contrato','IRRF': 'irrf', 'CIDE': 'cide', 'Moeda': 'moeda', 'Valor': 'valor_moeda_origem'}, inplace=True)
        cols_to_keep = ['data_pagamento', 'valor_principal_rs', 'numero_rm', 'correspondente', 'nd_recebida', 'iof', 'tx_contrato', 'irrf', 'cide', 'moeda', 'valor_moeda_origem']
        df = df[[col for col in cols_to_keep if col in df.columns]]
        numeric_cols = ['valor_principal_rs', 'iof', 'tx_contrato', 'irrf', 'cide', 'valor_moeda_origem']
        for col in numeric_cols: df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
        df['numero_rm'] = df['numero_rm'].fillna('Não informado').astype(str)
        df['nd_recebida'] = df['nd_recebida'].fillna('Não informada').astype(str)
        df['correspondente'] = df['correspondente'].fillna('Não informado')
        df['moeda'] = df['moeda'].fillna('N/A')
        df['data_pagamento'] = pd.to_datetime(df['data_pagamento'], errors='coerce', dayfirst=True)
        df['status'] = np.where(pd.notna(df['data_pagamento']), 'Pago', 'Pendente')
        df['valor_total_pago'] = df['valor_principal_rs'] + df['iof'] + df['tx_contrato'] + df['irrf'] + df['cide']
        return df[df['status'] == 'Pago'], df[df['status'] == 'Pendente']
    except FileNotFoundError:
        st.error(f"Erro: O arquivo não foi encontrado no caminho especificado: {file_path}")
        return None, None
    except Exception as e:
        st.error(f"Ocorreu um erro ao carregar o arquivo: {e}")
        return None, None

# --- Funções das Páginas ---
def display_visao_geral(df_pago):
    st.header("Visão Geral dos Pagamentos")
    st.subheader("Indicadores Chave de Performance (KPIs)")
    if not df_pago.empty:
        valor_total_geral_pago, maior_fatura_row, rm_maior_desembolso_group, corr_maior_pago_group = df_pago['valor_total_pago'].sum(), df_pago.loc[df_pago['valor_total_pago'].idxmax()], df_pago.groupby('numero_rm')['valor_total_pago'].sum(), df_pago.groupby('correspondente')['valor_total_pago'].sum()
        rm_maior_desembolso, corr_maior_pago = (rm_maior_desembolso_group.idxmax() if not rm_maior_desembolso_group.empty else "N/A"), (corr_maior_pago_group.idxmax() if not corr_maior_pago_group.empty else "N/A")
    else:
        valor_total_geral_pago, maior_fatura_row, rm_maior_desembolso, corr_maior_pago, rm_maior_desembolso_group, corr_maior_pago_group = 0, None, "N/A", "N/A", pd.Series(), pd.Series()
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Valor Total Pago", format_currency(valor_total_geral_pago))
    col2.metric("Maior Fatura Paga", format_currency(maior_fatura_row['valor_total_pago']) if maior_fatura_row is not None else "N/A", help=f"ND: {maior_fatura_row['nd_recebida']} | Corr: {maior_fatura_row['correspondente']}" if maior_fatura_row is not None else "")
    col3.metric("RM com Maior Desembolso", rm_maior_desembolso, help=f"Valor: {format_currency(rm_maior_desembolso_group.max()) if not rm_maior_desembolso_group.empty else 'R$ 0,00'}")
    col4.metric("Correspondente com Maior Desembolso", corr_maior_pago, help=f"Valor: {format_currency(corr_maior_pago_group.max()) if not corr_maior_pago_group.empty else 'R$ 0,00'}")
    st.markdown("<hr/>", unsafe_allow_html=True)
    st.subheader("Análises Visuais")
    st.markdown("#### Composição dos Pagamentos por Correspondente")
    show_all, legend_map = st.checkbox("Mostrar todos os correspondentes"), {'valor_principal_rs': 'Valor Principal R$', 'iof': 'IOF', 'irrf': 'IRRF', 'cide': 'CIDE', 'tx_contrato': 'Taxa do Contrato'}
    if not df_pago.empty:
        pag_por_corr, top_n = df_pago.groupby('correspondente')[list(legend_map.keys())].sum(), 0
        pag_por_corr['total'] = pag_por_corr.sum(axis=1)
        top_n = len(pag_por_corr) if show_all else 20
        data_to_plot = pag_por_corr.nlargest(top_n, 'total')
        data_to_plot_melted = data_to_plot.drop(columns=['total']).reset_index().melt(id_vars='correspondente', var_name='componente', value_name='valor')
        data_to_plot_melted = data_to_plot_melted.merge(data_to_plot[['total']], on='correspondente')
        data_to_plot_melted['percent'], data_to_plot_melted['componente'] = (data_to_plot_melted['valor'] / data_to_plot_melted['total']) * 100, data_to_plot_melted['componente'].map(legend_map)
        fig_bar_stacked = px.bar(data_to_plot_melted, x='valor', y='correspondente', color='componente', title=f"Top {top_n} Correspondentes", labels={'correspondente': '', 'valor': 'Valor Pago (R$)'}, orientation='h', barmode='stack', color_discrete_sequence=px.colors.sequential.Blues_r, custom_data=['total', 'percent'])
        fig_bar_stacked.update_traces(hovertemplate='<b>%{y}</b><br>Componente: %{data.name}<br>Valor: %{x:,.2f}<br>Total: %{customdata[0]:,.2f}<br>% do Total: %{customdata[1]:.2f}%<extra></extra>')
        fig_bar_stacked.update_layout(yaxis={'categoryorder':'total ascending'}, height=max(600, top_n * 25))
        st.plotly_chart(fig_bar_stacked, use_container_width=True)
    st.markdown("#### Proporção do Valor Total Pago por Correspondente")
    if not df_pago.empty:
        fig_treemap = px.treemap(df_pago, path=[px.Constant("Todos"), 'correspondente'], values='valor_total_pago', color='valor_total_pago', color_continuous_scale=px.colors.sequential.Blues)
        fig_treemap.update_traces(textinfo="label+percent root")
        st.plotly_chart(fig_treemap, use_container_width=True)

def display_rm_detalhado(df_pago, df_pendente):
    st.header("Análise Detalhada por RM")
    df_completo = pd.concat([df_pago, df_pendente])
    lista_rms = sorted(df_completo['numero_rm'].unique())
    
    numeric_rms = [int(r) for r in lista_rms if r.isdigit()]
    default_rm_val = [str(max(numeric_rms))] if numeric_rms else ([lista_rms[0]] if lista_rms else [])
    rm_selecionados = st.multiselect("Selecione um ou mais RMs para analisar:", lista_rms, default=default_rm_val)
    if rm_selecionados:
        df_rm = df_completo[df_completo['numero_rm'].isin(rm_selecionados)]
        st.subheader(f"Desembolso por Correspondente nos RMs Selecionados")
        df_rm_pago = df_rm[df_rm['status'] == 'Pago']
        if not df_rm_pago.empty:
            legend_map = {'valor_principal_rs': 'Valor Principal R$', 'iof': 'IOF', 'irrf': 'IRRF', 'cide': 'CIDE', 'tx_contrato': 'Taxa do Contrato'}
            pag_rm_corr = df_rm_pago.groupby('correspondente')[list(legend_map.keys())].sum()
            pag_rm_corr['total'] = pag_rm_corr.sum(axis=1)
            pag_rm_corr_melted = pag_rm_corr.drop(columns='total').reset_index().melt(id_vars='correspondente', var_name='componente', value_name='valor')
            pag_rm_corr_melted = pag_rm_corr_melted.merge(pag_rm_corr[['total']], on='correspondente')
            pag_rm_corr_melted['percent'], pag_rm_corr_melted['componente'] = (pag_rm_corr_melted['valor'] / pag_rm_corr_melted['total']) * 100, pag_rm_corr_melted['componente'].map(legend_map)
            fig_bar_rm = px.bar(pag_rm_corr_melted, x='valor', y='correspondente', color='componente', barmode='stack', orientation='h', labels={'correspondente': '', 'valor': 'Valor Pago (R$)'}, color_discrete_sequence=px.colors.sequential.Blues_r, custom_data=['total', 'percent'])
            fig_bar_rm.update_traces(hovertemplate='<b>%{y}</b><br>Componente: %{data.name}<br>Valor: %{x:,.2f}<br>Total: %{customdata[0]:,.2f}<br>% do Total: %{customdata[1]:.2f}%<extra></extra>')
            fig_bar_rm.update_layout(yaxis={'categoryorder':'total ascending'}, height=500)
            st.plotly_chart(fig_bar_rm, use_container_width=True)
        else: st.info("Não há pagamentos para os RMs selecionados.")
        st.subheader("Tabela de Lançamentos")
        cols_tabela = {'nd_recebida': 'Fatura (ND Recebida)', 'correspondente': 'Correspondente', 'iof': 'IOF (R$)', 'tx_contrato': 'TX Contrato (R$)', 'valor_principal_rs': 'Valor Principal (R$)', 'irrf': 'IRRF (R$)', 'cide': 'CIDE (R$)', 'valor_total_pago': 'Valor Total (R$)'}
        df_tabela = df_rm[[col for col in cols_tabela if col in df_rm]].copy()
        for col in cols_tabela:
            if '(R$)' in cols_tabela[col]: df_tabela[col] = df_tabela[col].apply(format_currency)
        st.dataframe(df_tabela.rename(columns=cols_tabela), use_container_width=True)

def display_correspondentes(df_pago, df_pendente):
    st.header("Análise por Correspondentes")
    col1, col2 = st.columns([0.6, 0.4])
    with col1:
        st.subheader("Ranking de Correspondentes")
        if not df_pago.empty:
            ranking = df_pago.groupby('correspondente')['valor_total_pago'].sum().sort_values(ascending=False).reset_index()
            ranking.rename(columns={'correspondente': 'Correspondente', 'valor_total_pago': 'Valor Pago em R$'}, inplace=True)
            ranking['Valor Pago em R$'] = ranking['Valor Pago em R$'].apply(format_currency)
            st.dataframe(ranking, use_container_width=True, height=400)
        else: st.info("Não há pagamentos para exibir o ranking.")
    with col2:
        st.subheader("Participação Percentual (Top 10)")
        if not df_pago.empty:
            ranking_raw = df_pago.groupby('correspondente')['valor_total_pago'].sum().nlargest(10)
            fig_pie = px.pie(ranking_raw, values='valor_total_pago', names=ranking_raw.index, hole=.4, color_discrete_sequence=px.colors.sequential.Blues_r)
            st.plotly_chart(fig_pie, use_container_width=True)
    st.markdown("<hr/>", unsafe_allow_html=True)
    st.subheader("Pendências por Correspondente")
    if not df_pendente.empty:
        pendencias = df_pendente.groupby(['correspondente', 'moeda'])['valor_moeda_origem'].agg(['count', 'sum']).sort_values(by='sum', ascending=False).reset_index()
        pendencias.rename(columns={'correspondente': 'Correspondente', 'moeda': 'Moeda', 'count': 'Qtd. Pendente', 'sum': 'Valor Pendente (Moeda de Origem)'}, inplace=True)
        pendencias['Valor Pendente (Moeda de Origem)'] = pendencias['Valor Pendente (Moeda de Origem)'].apply(format_foreign_currency)
        st.dataframe(pendencias, use_container_width=True)
    else: st.info("Não há faturas pendentes.")

# --- Interface Principal ---
st.title("🔵 Dashboard de Análise Financeira")

df_pago_original, df_pendente_original = load_data(FILE_PATH, SHEET_NAME, SHEET_PASSWORD)
if df_pago_original is None:
    st.error("Falha ao carregar os dados. Verifique as configurações de caminho e senha no script.")
    st.stop()
st.sidebar.success("Dados carregados com sucesso!")

st.sidebar.title("Filtros Globais")
df_completo_original = pd.concat([df_pago_original, df_pendente_original])
min_date, max_date = df_pago_original['data_pagamento'].min().date(), df_pago_original['data_pagamento'].max().date()
date_range = st.sidebar.date_input("Período de Pagamento:", value=(min_date, max_date), min_value=min_date, max_value=max_date)

correspondentes, select_all_corr = sorted(df_completo_original['correspondente'].unique()), st.sidebar.checkbox("Selecionar Todos os Correspondentes", value=True)
default_corr = correspondentes if select_all_corr else []
selected_correspondentes = st.sidebar.multiselect("Correspondente:", correspondentes, default=default_corr)

rms, select_all_rms = sorted(df_completo_original['numero_rm'].unique()), st.sidebar.checkbox("Selecionar Todos os RMs", value=True)
default_rms = rms if select_all_rms else []
selected_rms = st.sidebar.multiselect("Número do RM:", rms, default=default_rms)

df_pago_filtered, df_pendente_filtered = df_pago_original.copy(), df_pendente_original.copy()
if len(date_range) == 2:
    start_date, end_date = pd.to_datetime(date_range[0]), pd.to_datetime(date_range[1])
    df_pago_filtered = df_pago_filtered[df_pago_filtered['data_pagamento'].between(start_date, end_date)]
df_pago_filtered = df_pago_filtered[df_pago_filtered['correspondente'].isin(selected_correspondentes) & df_pago_filtered['numero_rm'].isin(selected_rms)]
df_pendente_filtered = df_pendente_filtered[df_pendente_filtered['correspondente'].isin(selected_correspondentes) & df_pendente_filtered['numero_rm'].isin(selected_rms)]

st.sidebar.title("Menu de Análise")
page = st.sidebar.radio("Escolha uma página:", ["Visão Geral", "RM Detalhado", "Correspondentes"])

if page == "Visão Geral":
    display_visao_geral(df_pago_filtered)
elif page == "RM Detalhado":
    display_rm_detalhado(df_pago_filtered, df_pendente_filtered)
elif page == "Correspondentes":
    display_correspondentes(df_pago_filtered, df_pendente_filtered)